# import the Flask class from the flask module
from flask import Flask, render_template, redirect, url_for, request, session, flash
from functools import wraps
from flask import Flask, render_template, request, send_file
import sqlite3 as sql
import csv
from openpyxl import Workbook
app = Flask(__name__)
with sql.connect('test.db') as conn:
    cur = conn.cursor()
print ("Opened database successfully")

# conn.execute('''CREATE TABLE players
#          (NAME           TEXT    NOT NULL,
#          score            INT     NOT NULL,
#          balls        INT,
#          sixes        INT );''')
# print ("Table created successfully")

# create the application object
app = Flask(__name__)

# config
app.secret_key = 'my precious'

# login required decorator
def login_required(f):
    @wraps(f)
    def wrap(*args, **kwargs):
        if 'logged_in' in session:
            return f(*args, **kwargs)
        else:
            flash('You need to login first.')
            return redirect(url_for('login'))
    return wrap

# use decorators to link the function to a url
@app.route('/')
@login_required
def home():
    return render_template('index.html')  # render a template
    # return "Hello, World!"  # return a string

@app.route('/welcome')
def welcome():
    return render_template('welcome.html')  # render a template

# route for handling the login page logic
@app.route('/login', methods=['GET', 'POST'])
def login():
    error = None
    if request.method == 'POST':
        if request.form['username'] != 'admin' or request.form['password'] != 'admin':
            error = 'Invalid Credentials. Please try again.'
        else:
            session['logged_in'] = True
            flash('You are logged in.')
            return render_template('student.html',error=error)
    return render_template('login.html', error=error)

@app.route('/logout')
@login_required
def logout():
    session.pop('logged_in', None)
    flash('You were logged out.')
    return redirect(url_for('welcome'))

@app.route('/student')
def details():
    return render_template('student.html')

@app.route('/addrec',methods = ['POST', 'GET'])
def addrec():
   if request.method == 'POST':
      try:
         nm = request.form['nm']
         score = request.form['score']
         balls = request.form['balls']
         sixes = request.form['sixes']

         # with sql.connect("database.db") as con:
         with sql.connect('test.db') as conn:
            cur = conn.cursor()
            cur.execute("INSERT INTO players (name,score,balls,sixes) VALUES (?,?,?,?)",(nm, score, balls, sixes) )

            conn.commit()
            msg = "Record successfully added"

            # ##
            fieldnames=['name','score','balls','sixes']

            with open('nameList.csv','a') as inFile:
                writer = csv.DictWriter(inFile, fieldnames=fieldnames)
                writer.writerow({'name': nm, 'score': score, 'balls':balls,'sixes':sixes})

            return render_template("result.html",msg=msg)

      except:
         conn.rollback()
         msg = "error in insert operation"

      finally:
         return render_template("result.html", msg=msg)
         conn.close()

@app.route('/list')
def list():
   conn = sql.connect("test.db")
   conn.row_factory = sql.Row

   cur = conn.cursor()
   cur.execute("select * from players")

   rows = cur.fetchall();
   return render_template("list.html",rows = rows)

@app.route('/getPlotCSV') # this is a job for GET, not POST
def plot_csv():

    # wb = Workbook()
    # db = 'test.db' # database where your table is stored
    # table = 'players' # table you want to save
    # SQL = "SELECT * FROM %s;" % table
    # with sql.connect('test.db') as conn:
    #    cur = conn.cursor()
    # cur.execute(SQL)
    # results = cur.fetchall()
    # ws = wb.create_sheet(0)
    # ws.title = "players"
    # ws.append(cur.column_names)
    # for row in results:
    #     ws.append(row)
    #
    # workbook_name = "test_workbook"
    # wb.save(workbook_name + ".xlsx")

    # user = '' # your username
    # passwd = '' # your password
    # host = '' # your host
    # db = 'test.db' # database where your table is stored
    # table = 'players' # table you want to save
    #
    # conn = sql.connect("test.db")
    # cur = conn.cursor()
    #
    # query = "SELECT * FROM %s;" % table
    # cur.execute(query)
    # workbook = Workbook('outfile.xlsx')
    # sheet = workbook.add_worksheet()
    # for r, row in enumerate(cur.fetchall()):
    #     for c, col in enumerate(row):
    #         sheet.write(r, c, col)
    #
    return send_file('nameList.csv',
                     mimetype='text/csv',
                     attachment_filename='Adjacency.csv',
                     as_attachment=True)





# start the server with the 'run()' method
if __name__ == '__main__':
    app.run(debug=True)
